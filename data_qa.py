import openpyxl as xl
from collections import Counter


def data_qa(workbook, worksheet):

    try:
        wb = xl.load_workbook(f'{workbook}.xlsx')
    except FileNotFoundError:
        error_message = f'File "{workbook}.xlsx" not found, please check filename and ensure it is in xlsx format.'
        return error_message

    try:
        sample_data = wb[f'{worksheet}']
    except KeyError:
        error_message = f'Sheet "{worksheet}" not found, please check that named sheet exists within "{workbook}.xlsx".'
        return error_message

    header_row = sample_data.min_row
    maxi_row = sample_data.max_row
    first_column = sample_data.min_column
    maxi_column = sample_data.max_column
    record_count = maxi_row - header_row  # do not need -1 here as header is not a record
    column_count = maxi_column - (first_column - 1)

    # print(header_row, maxi_row, first_column, maxi_column, record_count, column_count)

    result_blob = (
        f'\nAnalysed document "{workbook}.xlsx", sheet "{worksheet}" \n'
        f'Header Row: {header_row}, First Column: {first_column} \n'
        f'Total Records: {record_count}, Columns: {column_count} \n'
    )

    if record_count == 0:
        result_blob += 'Sheet is Empty.'
    else:
        result_blob += '\n'
        for columns in range(first_column, maxi_column+1):  # loops through each column
            name = str(sample_data.cell(header_row, columns).value)
            complete = 0  # used to count complete (not Null) records
            value = []  # used to count unique records
            types = []  # used to store data types

            for rows in range(header_row+1, maxi_row+1):  # loops through each row
                record = sample_data.cell(rows, columns).value  # selects cell value
                if record is not None:
                    complete += 1
                    rec_type = type(record).__name__  # method to return data type
                    types.append(rec_type)
                    if record not in value:
                        value.append(record)
            count_types = Counter(types).most_common()  # returns list of values and freq as [('str', 11), ('int', 1)]
            primary_type = count_types[0][0]

            all_types = ''  # used to store data types and counts

            for t in count_types:
                element = f'{t[0]} ({t[1]}), '
                all_types += element

            all_types = all_types[:-2]  # removes ', ' from last record

            value_count = len(value)

            result_blob += (
                f'{name}{' (!)' if len(count_types) > 1 else ''}: \n'
                f'{complete} total records ({format(complete/record_count, ".0%")} complete) \n'
                f'{value_count} unique records \n'
                f'{'Data type(s): ' + (all_types if len(count_types) > 1 else primary_type)} \n \n'
            )  # adds each column's data to the final output

    return result_blob

